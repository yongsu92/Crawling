import pandas as pd
import openpyxl
import csv
# no = [1,2,3]
# subjects = ['math','science','data']
#
# df = pd.DataFrame()
# df['SBJ_NUM'] = no
# df["NAME"] = subjects
# print(df)
# df.to_csv('test.csv',encoding='utf-8-sig',index=False)
# df.to_excel('test.xlsx',index=False)

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb['Sheet1']

contents = {}

for i in range(2,sheet.max_row+1):
    name = sheet.cell(row=i,column = 1).value
    email = sheet.cell(row=i,column = 2).value
    contents[name] = email
print(contents)



f = open('test.csv',encoding='utf-8')
f_csv = csv.reader(f)
for i in f_csv:
    print(i)
f.close()